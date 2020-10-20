from openpyxl import Workbook
import pymysql
conn = pymysql.connect(host='localhost', user='root', password='123456', port=3306, db='shopinfo', charset='utf8')
cursor = conn.cursor()


def save_shop_list():
    wb = Workbook(write_only=True)
    sheet = wb.create_sheet('local')
    row = ['storeId', 'shopName', 'eleId', 'monthSales', 'shopScore', 'wid', 'lat', 'lng']
    sheet.append(row)
    sql = 'select * from ele_shop'
    cursor.execute(sql)
    all_date = cursor.fetchall()
    for date in all_date:
        storeId = date[0]
        shopName = date[1]
        eleId = date[2]
        monthSales = date[3]
        shopScore = date[4]
        wid = date[5]
        lat = date[6]
        lng = date[7]
        row = [storeId, shopName, eleId, monthSales, shopScore, wid, lat, lng]
        sheet.append(row)
    file = 'D:\\top50\\shopList.xlsx'
    wb.save(file)


def save_shop_info():
    wb = Workbook(write_only=True)
    sheet = wb.create_sheet('local')
    row = ['storeId', 'shopName', 'monthSales', 'shopScore', 'address', 'activity', 'coupon', 'category', 'categoryIds', 'city']
    sheet.append(row)
    sql = 'select * from e_shop_info'
    cursor.execute(sql)
    all_date = cursor.fetchall()
    for date in all_date:
        storeId = date[0]
        shopName = date[1]
        monthSales = date[2]
        shopScore = date[3]
        address = date[4]
        activity = date[5]
        coupon = date[6]
        category = date[7]
        categoryIds = date[8]
        city = date[9]
        row = [storeId, shopName, monthSales, shopScore, address, activity, coupon, category, categoryIds, city]
        sheet.append(row)
    file = 'D:\\top50\\shopInfo.xlsx'
    wb.save(file)

# save_shop_info()